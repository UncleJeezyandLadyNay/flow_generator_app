#!/usr/bin/env python3
"""
generate.py — Generate an image or video on Google Flow using a saved session.

Requires a saved session from login.py (.session/flow_auth.json).

Usage:
    .venv/bin/python scripts/generate.py --prompt "A serene mountain lake at dawn" --type video
    .venv/bin/python scripts/generate.py --prompt "A photorealistic red fox" --type image
    .venv/bin/python scripts/generate.py --prompt "..." --type video --timeout 300 --out /path/to/dir
    .venv/bin/python scripts/generate.py --json-prompt prompt.json --type image
    .venv/bin/python scripts/generate.py --json-prompt '{"goal":"...","subject":[...]}' --type image

    # Pull prompt + filename from a spreadsheet row (1-based, row 1 = first data row):
    .venv/bin/python scripts/generate.py --xlsx prompts.xlsx --row 1 --type image --out ./output

Outputs the saved file path to stdout on success.
"""

import argparse
import json
import re
import sys
import time
import urllib.request
from datetime import datetime
from pathlib import Path

QUOTA_PATTERNS = re.compile(r"quota|rate.?limit|error.?253|253", re.I)

# ── paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent   # Google_Flow_Veo/
SKILL_ROOT = Path(__file__).resolve().parent.parent    # google-flow-generator/
SESSION_FILE = ROOT / ".session" / "flow_auth.json"
TMP_DIR = ROOT / ".tmp"
FLOW_URL = "https://labs.google/fx/tools/flow"


def ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


def log(msg: str) -> None:
    print(f"[{ts()}] {msg}", flush=True)


def json_to_prompt(data: dict) -> str:
    """
    Convert a Nano Banana 2 JSON prompt (12-field or 8-field schema) to a
    flat text string suitable for Google Flow's prompt input.

    Field priority follows the NB2 Five-Part Formula:
    [Cinematography] + [Subject] + [Action/Arrangement] + [Context] + [Style & Ambiance]
    Negative constraints are appended last.
    """
    parts = []

    # Goal — anchor/purpose sentence
    if data.get("goal"):
        parts.append(str(data["goal"]).strip())

    # Subject (12-field: array of traits | 8-field: array of descriptors)
    subject = data.get("subject")
    if subject:
        if isinstance(subject, list):
            parts.append(", ".join(str(s) for s in subject))
        else:
            parts.append(str(subject))

    # MadeOutOf — 8-field schema material/texture field
    made_of = data.get("MadeOutOf")
    if made_of:
        if isinstance(made_of, list):
            parts.append("made of " + ", ".join(str(m) for m in made_of))
        else:
            parts.append(f"made of {made_of}")

    # Arrangement — 8-field schema pose/placement
    if data.get("arrangement"):
        parts.append(str(data["arrangement"]))

    # Composition
    if data.get("composition"):
        parts.append(str(data["composition"]))

    # Context / Background (deduplicate if identical)
    context = data.get("context", "")
    background = data.get("background", "")
    if context:
        parts.append(str(context))
    if background and background != context:
        parts.append(str(background))

    # Style
    if data.get("style"):
        parts.append(str(data["style"]))

    # Lighting
    if data.get("lighting"):
        parts.append(str(data["lighting"]))

    # Color palette (12-field: color_palette | 8-field: ColorRestriction)
    palette = data.get("color_palette") or data.get("ColorRestriction")
    if palette:
        if isinstance(palette, list):
            parts.append("color palette: " + ", ".join(str(c) for c in palette))
        else:
            parts.append(f"color palette: {palette}")

    # Camera / lens (12-field: camera_or_lens | 8-field: camera)
    cam = data.get("camera_or_lens") or data.get("camera")
    if cam and isinstance(cam, dict):
        cam_parts = []
        if cam.get("type"):
            cam_parts.append(str(cam["type"]))
        lens = cam.get("lens") or cam.get("focal_length")
        if lens:
            cam_parts.append(f"{lens} lens")
        if cam.get("aperture"):
            cam_parts.append(str(cam["aperture"]))
        if cam.get("shutter_speed"):
            cam_parts.append(str(cam["shutter_speed"]))
        if cam.get("ISO"):
            cam_parts.append(f"ISO {cam['ISO']}")
        if cam.get("flash") and str(cam["flash"]).lower() != "none":
            cam_parts.append(f"{cam['flash']} flash")
        if cam_parts:
            parts.append(", ".join(cam_parts))

    # Mood
    if data.get("mood"):
        parts.append(f"mood: {data['mood']}")

    # Tags — 8-field schema keyword array
    tags = data.get("tags")
    if tags:
        if isinstance(tags, list):
            parts.extend(str(t) for t in tags)
        else:
            parts.append(str(tags))

    # Text space — only include if it reserves a position (skip "none")
    text_space = data.get("text_space", "")
    if text_space and str(text_space).lower() not in ("none", ""):
        parts.append(f"leave space for text at {text_space}")

    # Build base prompt
    prompt = ". ".join(p.strip().rstrip(".") for p in parts if p.strip())

    # Negative constraints appended as comma-separated list
    neg = data.get("negative_constraints")
    if neg:
        if isinstance(neg, list):
            prompt += ". " + ", ".join(str(n) for n in neg)
        else:
            prompt += f". {neg}"

    return prompt


def load_json_prompt(raw: str) -> str:
    """
    Accept either an inline JSON string or a file path.
    Returns a flat text prompt converted from the NB2 JSON schema.
    Attempts to auto-repair a missing closing `}` if the JSON is otherwise complete.
    """
    raw = raw.strip()
    if raw.startswith("{"):
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            # Auto-repair: append missing closing brace if the object looks complete
            repaired = raw.rstrip() + "\n}"
            try:
                data = json.loads(repaired)
                log("WARNING: JSON prompt was missing closing `}` — auto-repaired.")
            except json.JSONDecodeError as e:
                raise ValueError(f"Invalid JSON prompt (auto-repair failed): {e}") from e
    else:
        with open(raw) as f:
            data = json.load(f)
    return json_to_prompt(data)


def read_xlsx_row(xlsx_path: str, row_index: int) -> tuple[str, str]:
    """
    Read a single row from an xlsx spreadsheet and return (prompt_text, filename).

    Args:
        xlsx_path:  Path to the .xlsx file.
        row_index:  1-based row number in the data (row 1 = first data row after header).

    Columns expected (case-insensitive):
        Prompt   — JSON string matching the Nano Banana 2 schema; compiled to flat text.
        Filename — Output filename stem (no extension needed).

    Raises:
        FileNotFoundError if the xlsx doesn't exist.
        ValueError if required columns are missing or the row index is out of range.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed.", file=sys.stderr)
        print("Run:  .venv/bin/pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"xlsx file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Build column index map from header row (row 1 in the sheet)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {
        str(cell).strip().lower(): idx
        for idx, cell in enumerate(header_row)
        if cell is not None
    }

    required = {"prompt", "filename"}
    missing = required - col_map.keys()
    if missing:
        raise ValueError(
            f"xlsx is missing required column(s): {', '.join(sorted(missing))}. "
            f"Found: {', '.join(col_map.keys())}"
        )

    prompt_col  = col_map["prompt"]
    filename_col = col_map["filename"]

    # Data rows start at sheet row 2 (row 1 is the header)
    sheet_row = row_index + 1  # row_index is 1-based data row → sheet row
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    if row_index < 1 or row_index > len(all_rows):
        raise ValueError(
            f"Row index {row_index} is out of range. "
            f"Spreadsheet has {len(all_rows)} data row(s) (valid range: 1–{len(all_rows)})."
        )

    row = all_rows[row_index - 1]  # convert to 0-based list index

    raw_prompt = row[prompt_col]
    filename   = row[filename_col]

    if raw_prompt is None:
        raise ValueError(f"Prompt cell is empty at row {row_index}.")
    if filename is None:
        raise ValueError(f"Filename cell is empty at row {row_index}.")

    # Prompt column may contain either a Nano Banana 2 JSON object or plain text.
    # Excel sometimes wraps cell content in outer quotes: `"{ ... }"` — strip them first.
    raw_prompt_str = str(raw_prompt).strip()
    if raw_prompt_str.startswith('"') and raw_prompt_str.endswith('"'):
        raw_prompt_str = raw_prompt_str[1:-1]

    # If it looks like JSON, compile it; otherwise use the text directly.
    if raw_prompt_str.lstrip().startswith("{"):
        prompt_text = load_json_prompt(raw_prompt_str)
    else:
        prompt_text = raw_prompt_str
    filename_stem = Path(str(filename).strip()).stem  # strip extension if user included one

    return prompt_text, filename_stem


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate media on Google Flow")
    parser.add_argument("--prompt", required=False, default=None, help="Text prompt to generate from")
    parser.add_argument(
        "--json-prompt",
        dest="json_prompt",
        default=None,
        help="Nano Banana 2 JSON prompt: path to a .json file OR inline JSON string",
    )
    parser.add_argument(
        "--type",
        choices=["image", "video"],
        default="video",
        help="Type of media to generate (default: video)",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=TMP_DIR,
        help="Output directory (default: .tmp/)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=None,
        help="Max seconds to wait for generation (default: 600 for video, 300 for image)",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        default=False,
        help="Run headless (default: False). Google Flow detects and blocks standard headless "
             "Chrome — prefer the default headed mode. Use --headless to try Chrome's new "
             "headless mode, which is harder to detect.",
    )
    parser.add_argument("--no-headless", dest="headless", action="store_false")
    parser.add_argument(
        "--retries",
        type=int,
        default=3,
        help="Max retries on quota errors (default: 3)",
    )
    parser.add_argument(
        "--retry-delay",
        type=int,
        default=60,
        help="Seconds to wait between retries (default: 60)",
    )
    parser.add_argument(
        "--name",
        type=str,
        default=None,
        help="Output filename stem (no extension). Default: timestamp_type.",
    )
    # ── xlsx source ────────────────────────────────────────────────────────────
    parser.add_argument(
        "--xlsx",
        type=str,
        default=None,
        help="Path to an .xlsx file containing 'Prompt' (JSON) and 'Filename' columns.",
    )
    parser.add_argument(
        "--row",
        type=int,
        default=None,
        help="1-based row index in the xlsx data (row 1 = first data row after header).",
    )
    return parser.parse_args()


def check_session() -> dict:
    """Load and return saved session state, exit cleanly if missing."""
    if not SESSION_FILE.exists():
        print("ERROR: No saved session found.", file=sys.stderr)
        print(f"       Expected: {SESSION_FILE}", file=sys.stderr)
        print("       Run login.py first to authenticate:", file=sys.stderr)
        print("         .venv/bin/python scripts/login.py", file=sys.stderr)
        sys.exit(1)
    with open(SESSION_FILE) as f:
        return json.load(f)


def detect_session_expired(page) -> bool:
    """Return True if the page looks like a Google login/redirect page."""
    url = page.url
    return (
        "accounts.google.com" in url
        or "signin" in url.lower()
        or "/login" in url.lower()
    )


def find_prompt_input(page):
    """
    Find the Flow editor prompt input.
    The editor uses a contenteditable div with role=textbox and placeholder
    'What do you want to create?' — not a <textarea>.
    Returns the locator or None.
    """
    candidates = [
        # Flow editor: contenteditable div with role=textbox
        page.locator("div[role='textbox'][contenteditable='true']"),
        page.locator("[contenteditable='true'][role='textbox']"),
        # Fallback: any visible contenteditable
        page.locator("[contenteditable='true']:visible"),
        # Legacy: aria/placeholder based
        page.get_by_role("textbox", name=re.compile(r"prompt|describe|write|create", re.I)),
        page.get_by_placeholder(re.compile(r"prompt|describe|type|create", re.I)),
        page.locator("textarea:visible"),
    ]
    for locator in candidates:
        try:
            if locator.count() > 0 and locator.first.is_visible():
                return locator.first
        except Exception:
            continue
    return None


def find_generate_button(page):
    """
    Find the Flow editor generate/submit button.
    The editor has two 'Create' buttons: 'add_2Create' (add media) and
    'arrow_forwardCreate' (submit/generate). We want the arrow_forward one.
    Returns the locator or None.
    """
    # Prefer the arrow_forward submit button over the add_2 add button
    # Both contain 'Create' in text; the submit one has 'arrow_forward' icon text
    try:
        all_btns = page.locator("button")
        for i in range(all_btns.count()):
            btn = all_btns.nth(i)
            try:
                txt = (btn.text_content() or "").strip()
                if "arrow_forward" in txt and "create" in txt.lower() and btn.is_visible():
                    return btn
            except Exception:
                continue
    except Exception:
        pass

    # Fallback: any visible generate/create/run button
    candidates = [
        page.locator("button[aria-label*='generate' i]"),
        page.locator("button[aria-label*='submit' i]"),
        page.get_by_role("button", name=re.compile(r"^generate$|^create$|^run$", re.I)),
        page.locator("[data-testid*='generate']"),
        page.locator("[data-testid*='submit']"),
    ]
    for locator in candidates:
        try:
            if locator.count() > 0 and locator.first.is_visible():
                return locator.first
        except Exception:
            continue
    return None


def find_media_type_selector(page, media_type: str) -> bool:
    """
    Attempt to switch the UI to 'image' or 'video' mode.
    Returns True if successfully clicked, False if not found (non-fatal).
    """
    # Try tab/button selectors for image vs video
    pattern = re.compile(media_type, re.I)
    candidates = [
        page.get_by_role("tab", name=pattern),
        page.get_by_role("button", name=pattern),
        page.get_by_role("radio", name=pattern),
        page.locator(f"[aria-label*='{media_type}' i]"),
        page.locator(f"[data-value='{media_type}']"),
    ]
    for locator in candidates:
        try:
            if locator.count() > 0:
                locator.first.click()
                log(f"Selected media type: {media_type}")
                return True
        except Exception:
            continue
    log(f"WARNING: Could not find explicit '{media_type}' type selector — proceeding with default")
    return False


def select_video_model(page, model_name: str = "Veo 3.1 - Lite") -> bool:
    """
    Switch Flow to video mode and select a Veo model.

    Flow's toolbar button opens a settings panel containing:
      - Image / Video tabs (must click Video tab first)
      - A model dropdown (arrow_drop_down) that shows Veo options after switching to Video
    Returns True if a Veo model was selected, False if not found (non-fatal).
    """
    # Step 1: open the settings panel via the toolbar model button
    opener_candidates = [
        page.get_by_role("button", name=re.compile(r"veo", re.I)),
        page.locator("button", has_text=re.compile(r"veo", re.I)),
        page.locator("button", has_text=re.compile(r"nano.?banana|🍌", re.I)),
        page.get_by_role("button", name=re.compile(r"nano.?banana|🍌", re.I)),
        page.locator("[aria-label*='model' i]"),
        page.locator("[data-testid*='model' i]"),
    ]
    opened = False
    for loc in opener_candidates:
        try:
            if loc.count() > 0 and loc.first.is_visible():
                loc.first.click()
                page.wait_for_timeout(800)
                opened = True
                log(f"Opened settings panel ({(loc.first.text_content() or '').strip()[:50]!r})")
                break
        except Exception:
            continue

    if not opened:
        log(f"WARNING: Could not open settings panel for '{model_name}' — using default")
        return False

    # Step 2: click the Video tab within the panel (Image/Video tabs live inside the panel)
    video_tab_candidates = [
        page.get_by_role("tab", name=re.compile(r"video", re.I)),
        page.locator("[role='tab']", has_text=re.compile(r"video", re.I)),
        page.locator("button[role='tab']", has_text=re.compile(r"video", re.I)),
    ]
    for loc in video_tab_candidates:
        try:
            if loc.count() > 0 and loc.first.is_visible():
                loc.first.click()
                page.wait_for_timeout(800)
                log("Switched to Video tab")
                break
        except Exception:
            continue

    # Step 3: open the model dropdown (button text contains 'arrow_drop_down' icon name)
    dropdown_candidates = [
        page.locator("button", has_text=re.compile(r"arrow_drop_down", re.I)),
        page.locator("[aria-haspopup='listbox']"),
        page.locator("[aria-haspopup='menu']"),
    ]
    for loc in dropdown_candidates:
        try:
            if loc.count() > 0 and loc.first.is_visible():
                loc.first.click()
                page.wait_for_timeout(600)
                log("Opened model dropdown")
                break
        except Exception:
            continue

    # Step 4: pick the Veo option; try progressively looser matches
    name_variants = [model_name, "Veo 3.1 Lite", "Veo 3.1 - Lite", "Lite", "Veo 3.1", "Veo"]
    for variant in name_variants:
        option_candidates = [
            page.get_by_role("option",   name=re.compile(re.escape(variant), re.I)),
            page.get_by_role("menuitem", name=re.compile(re.escape(variant), re.I)),
            page.get_by_role("radio",    name=re.compile(re.escape(variant), re.I)),
            page.locator("button",       has_text=re.compile(re.escape(variant), re.I)),
            page.locator(f"[aria-label*='{variant}' i]"),
            page.get_by_text(re.compile(re.escape(variant), re.I)),
        ]
        for loc in option_candidates:
            try:
                if loc.count() > 0 and loc.first.is_visible():
                    loc.first.click()
                    log(f"Selected model: {variant}")
                    page.wait_for_timeout(600)
                    return True
            except Exception:
                continue

    log(f"WARNING: Veo model option not found — using default (Video tab was still activated)")
    return False


def set_variation_count(page, count: int = 1) -> bool:
    """
    Set the number of video variations to `count`.
    Variation buttons (1x, x2, x3, x4) are directly visible in the settings panel.
    Verifies aria-selected after clicking and retries up to 3 times since the panel
    can still be updating after a model switch.
    Returns True if confirmed selected, False if not found (non-fatal).
    """
    target = str(count)
    # Flow uses "1x" for 1 variation and "x2"/"x3"/"x4" for others
    btn_pattern = re.compile(rf"^x?{target}x?$", re.I)

    for attempt in range(3):
        if attempt > 0:
            page.wait_for_timeout(600)  # panel may still be settling after model switch

        # Find the target button
        btn = None
        for loc in [
            page.locator("button", has_text=btn_pattern),
            page.get_by_role("button", name=btn_pattern),
        ]:
            try:
                if loc.count() > 0 and loc.first.is_visible():
                    btn = loc.first
                    break
            except Exception:
                continue

        if btn is None:
            continue

        btn.click()
        page.wait_for_timeout(400)

        # Verify it actually took — check aria-selected on the clicked button
        try:
            selected = btn.get_attribute("aria-selected")
            if selected == "true":
                log(f"Set variation count to {count} (confirmed)")
                return True
        except Exception:
            pass

        log(f"  Variation click attempt {attempt + 1} — not confirmed yet, retrying...")

    log(f"WARNING: Could not confirm variation count = {count} — using whatever Flow defaulted to")
    return False


def snapshot_image_urls(page) -> set:
    """Return the set of img src URLs currently on the page."""
    try:
        srcs = page.evaluate("""() =>
            [...document.querySelectorAll('img[src]')].map(i => i.src)
        """)
        return set(srcs or [])
    except Exception:
        return set()


# Patterns that indicate a generated media response (not a UI asset)
# CDN domains that serve actual generated media from Flow
MEDIA_CDN_PATTERNS = re.compile(
    r"flow-content\.google|"
    r"storage\.googleapis\.com.*\.(mp4|webm|png|jpg|jpeg|webp)|"
    r"lh3\.googleusercontent\.com",
    re.I,
)
GENERATED_URL_PATTERNS = re.compile(
    r"\.(mp4|webm|png|jpg|jpeg|gif|webp)(\?|$)|"
    r"generated|generate|output|result|render|artifact",
    re.I,
)
# URLs to never use as a result
SKIP_URL_PATTERNS = re.compile(
    r"gstatic\.com|/banner|/icon|/logo|favicon|sprite|placeholder|"
    r"fonts\.google|analytics|gtag|recaptcha|"
    r"labs\.google/fx/images/|perlin",  # Flow UI background textures
    re.I,
)
# Only skip aisandbox JSON endpoints — not actual media responses from that domain
SKIP_JSON_ONLY_PATTERNS = re.compile(r"aisandbox-pa\.googleapis\.com", re.I)


def install_response_interceptor(page, collected: list, media_type: str,
                                  video_urls: list | None = None) -> None:
    """
    Listen for network responses that look like generated media.
    Appends matching URLs to `collected`.
    Confirmed video/* responses are also appended to `video_urls` (if provided).

    Five capture paths:
    1. video/* content-type — always captured regardless of domain.
    2. Any flow-content.google URL in the response URL itself.
    3. Content-type match (image/*) for direct media responses.
    4. URL pattern match (.mp4/.webm/.png etc, or keywords like "generated").
    5. JSON body scan — Flow delivers generated media URLs inside JSON API
       responses. Parse the body and extract embedded media URLs.
    """
    FLOW_URL_IN_BODY = re.compile(r'https://flow-content\.google/[^\s"\']+', re.I)
    MP4_IN_BODY      = re.compile(r'https://[^\s"\'\\]+\.(?:mp4|webm)[^\s"\'\\]*', re.I)

    def on_response(response):
        try:
            url = response.url
            ct = response.headers.get("content-type", "")

            # Path 1: always capture any video/* response — never filter by domain.
            # Flow may serve video files from aisandbox-pa.googleapis.com or other
            # domains not in MEDIA_CDN_PATTERNS.
            if "video/" in ct:
                if url not in collected:
                    log(f"  [net] Video captured: {url[:100]}  ct={ct[:40]}")
                    collected.append(url)
                if video_urls is not None and url not in video_urls:
                    video_urls.append(url)
                return

            # Path 2–4: domain-filtered image / CDN / extension matching.
            # aisandbox-pa.googleapis.com normally serves JSON API responses; skip those.
            # But if it returns image/* content, it's a real generated image — allow it.
            is_media_ct = "image/" in ct  # video/* is already handled in Path 1
            skip        = SKIP_URL_PATTERNS.search(url) or (
                SKIP_JSON_ONLY_PATTERNS.search(url) and not is_media_ct
            )
            if not skip:
                is_flow_cdn  = MEDIA_CDN_PATTERNS.search(url)
                is_media_url = GENERATED_URL_PATTERNS.search(url)
                if (is_flow_cdn or is_media_ct or is_media_url) and url not in collected:
                    log(f"  [net] Candidate: {url[:100]}  ct={ct[:40]}")
                    collected.append(url)

            # Path 5: scan JSON bodies for embedded media URLs
            if "json" in ct:
                try:
                    body = response.text()
                    for match in FLOW_URL_IN_BODY.finditer(body):
                        found_url = match.group(0).rstrip('\\",}]')
                        if found_url not in collected:
                            log(f"  [net] Found in JSON body: {found_url[:100]}")
                            collected.append(found_url)
                    for match in MP4_IN_BODY.finditer(body):
                        found_url = match.group(0).rstrip('\\",}]')
                        if found_url not in collected:
                            log(f"  [net] Found mp4 in JSON body: {found_url[:100]}")
                            collected.append(found_url)
                except Exception:
                    pass
        except Exception:
            pass

    page.on("response", on_response)


FLOW_BASE = "https://labs.google"


def make_absolute(url: str) -> str:
    """Resolve a relative /path URL to an absolute https://labs.google/path URL."""
    if url and url.startswith("/"):
        return FLOW_BASE + url
    return url


def wait_for_result(page, media_type: str, timeout_sec: int,
                    pre_urls: set | None = None,
                    intercepted: list | None = None,
                    video_urls: list | None = None) -> str | None:
    """
    Poll the DOM for a completed generation result.
    Returns a URL to the generated media, or None on timeout.

    Strategy: wait for a new <video> src or <img> src (not a placeholder/logo)
    to appear in the result area.
    """
    log(f"Waiting for {media_type} generation (timeout: {timeout_sec}s)...")
    deadline = time.time() + timeout_sec
    poll_interval = 5
    min_wait = 15  # don't check for errors in the first 15s (avoids false positives)
    min_result_wait = 15  # don't trust intercepted URLs until generation has had time to start

    while time.time() < deadline:
        time.sleep(poll_interval)
        elapsed = int(time.time() - (deadline - timeout_sec))
        log(f"  Still waiting... ({elapsed}s elapsed)")

        elapsed_so_far = time.time() - (deadline - timeout_sec)


        # 0th pass (video only): confirmed video/* network responses — highest confidence,
        # captured from any CDN domain (e.g. aisandbox-pa.googleapis.com, flow-content.google).
        if media_type == "video" and video_urls and elapsed_so_far >= min_result_wait:
            if video_urls:
                log(f"  [net] Returning confirmed video URL: {video_urls[0][:80]}")
                return video_urls[0]

        # Primary: check network-intercepted URLs — prefer known media CDNs first
        if intercepted and elapsed_so_far >= min_result_wait:
            # 1st pass: flow-content.google CDN URLs, filtered to the right media type
            for url in list(intercepted):
                if MEDIA_CDN_PATTERNS.search(url) and not SKIP_URL_PATTERNS.search(url):
                    # Skip wrong-type CDN paths (e.g. /image/ when expecting video)
                    if media_type == "video" and re.search(r"flow-content\.google/image/", url, re.I):
                        continue
                    if media_type == "image" and re.search(r"flow-content\.google/video/", url, re.I):
                        continue
                    return url
            # 2nd pass: URL with a media extension matching the expected type
            for url in list(intercepted):
                if not SKIP_URL_PATTERNS.search(url):
                    if media_type == "video" and re.search(r"\.(mp4|webm)", url, re.I):
                        return url
                    if media_type == "image" and re.search(r"\.(png|jpg|jpeg|webp|gif)", url, re.I):
                        return url
            # 3rd pass (image only): any intercepted URL from a googleapis.com domain
            # — aisandbox-pa may serve generated images with opaque URL paths (no extension).
            if media_type == "image":
                for url in list(intercepted):
                    if not SKIP_URL_PATTERNS.search(url) and "googleapis.com" in url:
                        return url

        if media_type == "video":
            # Primary: check whether any <video> element has loaded real video data.
            # video.duration > 0 means the browser decoded actual video frames, not
            # just a thumbnail poster — this is the most reliable "generation done" signal.
            try:
                video_info = page.evaluate("""() => {
                    const videos = [...document.querySelectorAll('video')];
                    for (const v of videos) {
                        if ((v.duration > 0 || (v.readyState >= 2 && v.videoWidth > 0)) && v.src) {
                            return {src: v.src, currentSrc: v.currentSrc || v.src,
                                    duration: v.duration, readyState: v.readyState,
                                    videoWidth: v.videoWidth};
                        }
                    }
                    return null;
                }""")
                if video_info and (video_info.get("duration", 0) > 0 or
                                   video_info.get("readyState", 0) >= 2):
                    vsrc = video_info.get("currentSrc") or video_info.get("src", "")
                    log(f"  [dom] Video loaded (duration={video_info.get('duration',0):.1f}s "
                        f"readyState={video_info.get('readyState')} width={video_info.get('videoWidth')})")
                    if vsrc:
                        vsrc = make_absolute(vsrc)
                        # For media.getMediaUrlRedirect URLs, send Accept: video/* so the
                        # server returns the video redirect instead of the jpeg thumbnail.
                        video_accept = {"Accept": "video/mp4,video/webm,video/*;q=0.9,*/*;q=0.8"}
                        if "media.getMediaUrlRedirect" in vsrc:
                            try:
                                resp = page.context.request.get(vsrc, headers=video_accept)
                                ct = resp.headers.get("content-type", "")
                                final = resp.url
                                if "video" in ct or re.search(r"\.(mp4|webm)", final, re.I):
                                    log(f"  [dom] Redirect → video: {final[:80]}")
                                    return final
                                log(f"  [dom] Video ready (browser) but redirect ct={ct!r} — "
                                    f"returning redirect URL for download")
                                return vsrc
                            except Exception as exc:
                                log(f"  [dom] Redirect follow failed ({exc}) — using src directly")
                                return vsrc
                        return vsrc
            except Exception:
                pass

            # Secondary: find <video src="*media.getMediaUrlRedirect*"> and follow with
            # correct Accept header so the server serves the video instead of the thumbnail.
            try:
                video_srcs = page.evaluate("""() =>
                    [...document.querySelectorAll('video[src*="media.getMediaUrlRedirect"]')]
                        .map(v => v.src).filter(s => s.length > 0)
                """)
                video_accept = {"Accept": "video/mp4,video/webm,video/*;q=0.9,*/*;q=0.8"}
                for redirect_url in (video_srcs or []):
                    redirect_url = make_absolute(redirect_url)
                    try:
                        resp = page.context.request.get(redirect_url, headers=video_accept)
                        ct = resp.headers.get("content-type", "")
                        final_url = resp.url
                        if "video" in ct or re.search(r"\.(mp4|webm)", final_url, re.I):
                            log(f"  [dom] Video redirect resolved: {final_url[:80]}")
                            return final_url
                        if elapsed_so_far > 30 and elapsed_so_far % 30 < poll_interval:
                            log(f"  [dom] Redirect still serving {ct!r} — video not ready yet")
                    except Exception:
                        pass
            except Exception:
                pass

            # Fallback: blob: src or <video source> or download link
            for selector in ("video[src^='blob:']", "video source[src]"):
                try:
                    els = page.locator(selector)
                    if els.count() > 0:
                        src = els.first.get_attribute("src")
                        if src:
                            return src
                except Exception:
                    pass
            try:
                download_links = page.locator("a[download]")
                if download_links.count() > 0:
                    href = download_links.first.get_attribute("href")
                    if href:
                        return href
            except Exception:
                pass

        elif media_type == "image":
            try:
                imgs = page.locator("img[src]")
                all_srcs = []
                for i in range(imgs.count()):
                    src = imgs.nth(i).get_attribute("src") or ""
                    all_srcs.append(src)
                    # Always check media.getMediaUrlRedirect — these are generated images
                    if "media.getMediaUrlRedirect" in src:
                        abs_src = make_absolute(src)
                        if not pre_urls or src not in pre_urls:
                            log(f"  [dom] Generated image URL: {abs_src[:80]}")
                            return abs_src
                        # URL is in pre_urls — follow redirect to check if it's a real image
                        try:
                            resp = page.context.request.get(abs_src)
                            if "image/" in resp.headers.get("content-type", "") and resp.ok:
                                log(f"  [dom] Image redirect resolved: {resp.url[:80]}")
                                return resp.url
                        except Exception:
                            pass
                        continue
                    # Skip URLs that existed before we clicked generate
                    if pre_urls and src in pre_urls:
                        continue
                    if not src.startswith("https://"):
                        continue
                    low = src.lower()
                    if any(d in low for d in ["gstatic.com", "logo", "icon", "banner",
                                              "favicon", "sprite", "placeholder", "/assets/"]):
                        continue
                    if len(src) < 50:
                        continue
                    return make_absolute(src)
            except Exception:
                pass

        # Check for error state (only after min_wait to avoid false positives).
        # Use role=alert only — aria-label*='error' is too broad and catches brand elements.
        elapsed = time.time() - (deadline - timeout_sec)
        if elapsed >= min_wait:
            try:
                error = page.locator("[role='alert']")
                if error.count() > 0:
                    msg = (error.first.inner_text() or "").strip()
                    # Require error-like keywords to avoid matching live regions / toasts
                    if msg and re.search(r"error|fail|quota|limit|unavail|sorry|wrong", msg, re.I):
                        log(f"ERROR: UI reported an error: {msg}")
                        if QUOTA_PATTERNS.search(msg):
                            return "QUOTA_ERROR"
                        return None
            except Exception:
                pass

    log("ERROR: Timed out waiting for generation result.")
    if video_urls:
        log(f"  Confirmed video URL(s) intercepted ({len(video_urls)}):")
        for u in video_urls:
            log(f"    {u[:120]}")
    if intercepted:
        log(f"  All intercepted URL(s) ({len(intercepted)}):")
        for u in intercepted:
            log(f"    {u[:120]}")
    else:
        log("  No URLs were intercepted during wait.")
    return None


def download_result(url: str, media_type: str, out_dir: Path, page, name: str | None = None) -> Path:
    """
    Download the generated media from `url` to `out_dir`.
    Uses Playwright's APIRequestContext (bypasses CORS, carries browser cookies).
    Returns the saved file Path.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(name).stem if name else f"{timestamp}_{media_type}"

    if url.startswith("blob:"):
        # blob: URLs must be read from within the page JS context
        log("Extracting blob: URL via JavaScript...")
        import base64
        data_b64 = page.evaluate(
            """async (url) => {
                const resp = await fetch(url);
                const blob = await resp.blob();
                return new Promise((resolve) => {
                    const reader = new FileReader();
                    reader.onload = () => resolve(reader.result.split(',')[1]);
                    reader.readAsDataURL(blob);
                });
            }""",
            url,
        )
        ext = "mp4" if media_type == "video" else "png"
        out_path = out_dir / f"{stem}.{ext}"
        out_path.write_bytes(base64.b64decode(data_b64))
    else:
        # Use Playwright's request context — bypasses CORS, carries session cookies.
        # For media.getMediaUrlRedirect URLs, send Accept: video/* so the server
        # redirects to the video file rather than the jpeg thumbnail.
        extra_headers = {}
        if media_type == "video" and "media.getMediaUrlRedirect" in url:
            extra_headers = {"Accept": "video/mp4,video/webm,video/*;q=0.9,*/*;q=0.8"}
        log(f"Downloading via Playwright request: {url[:80]}...")
        response = page.context.request.get(url, headers=extra_headers)
        if not response.ok:
            raise RuntimeError(f"Download failed: HTTP {response.status} for {url[:80]}")
        # Determine extension from Content-Type header
        ct = response.headers.get("content-type", "")
        ext_map = {
            "image/png": "png", "image/jpeg": "jpg", "image/webp": "webp",
            "image/gif": "gif", "video/mp4": "mp4", "video/webm": "webm",
        }
        ext = next((v for k, v in ext_map.items() if k in ct), "mp4" if media_type == "video" else "png")
        out_path = out_dir / f"{stem}.{ext}"
        out_path.write_bytes(response.body())

    return out_path


def main():
    args = parse_args()

    if args.timeout is None:
        args.timeout = 600 if args.type == "video" else 300

    # ── Resolve prompt source (xlsx > json-prompt > prompt) ───────────────────
    if args.xlsx:
        # xlsx overrides all other prompt sources
        if args.prompt or args.json_prompt:
            print("ERROR: Use --xlsx OR --prompt/--json-prompt, not both.", file=sys.stderr)
            sys.exit(1)
        if args.row is None:
            print("ERROR: --row is required when using --xlsx.", file=sys.stderr)
            sys.exit(1)
        try:
            args.prompt, xlsx_filename = read_xlsx_row(args.xlsx, args.row)
        except (FileNotFoundError, ValueError) as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(1)
        # xlsx filename wins unless user also passed --name explicitly
        if args.name is None:
            args.name = xlsx_filename
        log(f"xlsx row {args.row}: filename={args.name!r}")
        log(f"xlsx prompt compiled: {args.prompt[:120]}{'...' if len(args.prompt) > 120 else ''}")
    elif args.json_prompt and args.prompt:
        print("ERROR: Use --prompt OR --json-prompt, not both.", file=sys.stderr)
        sys.exit(1)
    elif not args.json_prompt and not args.prompt:
        print("ERROR: One of --prompt, --json-prompt, or --xlsx/--row is required.", file=sys.stderr)
        sys.exit(1)
    elif args.json_prompt:
        args.prompt = load_json_prompt(args.json_prompt)
        log(f"JSON prompt compiled: {args.prompt[:120]}{'...' if len(args.prompt) > 120 else ''}")

    storage_state = check_session()

    args.out.mkdir(parents=True, exist_ok=True)

    log(f"Starting Google Flow generator")
    log(f"  Prompt : {args.prompt[:80]}{'...' if len(args.prompt) > 80 else ''}")
    log(f"  Type   : {args.type}")
    log(f"  Output : {args.out}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("ERROR: Playwright is not installed.", file=sys.stderr)
        print("Run:  .venv/bin/pip install playwright && .venv/bin/playwright install chromium", file=sys.stderr)
        sys.exit(1)

    # Google Flow detects and blocks standard headless Chrome (--headless flag).
    # --headless=new uses Chrome's modern headless mode which shares the same renderer
    # as headed Chrome, making it much harder to detect as automation.
    # We pass headless=False to Playwright (so it doesn't add the old --headless flag)
    # and add --headless=new ourselves when the user requests headless mode.
    launch_headless = False  # always False — we control headless via launch_args
    launch_args = [
        "--disable-blink-features=AutomationControlled",
        *( ["--headless=new"] if args.headless else ["--start-maximized"] ),
    ]

    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(
                headless=launch_headless,
                channel="chrome",
                args=launch_args,
            )
        except Exception:
            browser = p.chromium.launch(
                headless=launch_headless,
                args=launch_args,
            )

        # In --headless=new mode, Chrome has no display server.
        # Set an explicit viewport so the page renders at a predictable size.
        viewport = {"width": 1920, "height": 1080} if args.headless else None
        context = browser.new_context(
            storage_state=storage_state,
            viewport=viewport,
            java_script_enabled=True,
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()

        # ── Navigate ───────────────────────────────────────────────────────────
        log(f"Navigating to {FLOW_URL} ...")
        page.goto(FLOW_URL, wait_until="networkidle", timeout=30_000)

        if detect_session_expired(page):
            log("ERROR: Session expired. Run login.py to re-authenticate:")
            log("  .venv/bin/python scripts/login.py")
            context.close()
            browser.close()
            sys.exit(1)

        log(f"Page loaded: {page.url}")

        # The Flow URL loads the project gallery. Click "New project" to open the editor.
        # If we're already on a project page (previous redirect), skip this step.
        if "/project/" not in page.url:
            log("Looking for 'New project' button in project gallery...")
            try:
                # Use has_text for partial text match — button text is "add_2New project"
                new_proj_candidates = [
                    page.locator("button", has_text=re.compile(r"new project", re.I)),
                    page.get_by_role("button", name=re.compile(r"new project", re.I)),
                ]
                clicked = False
                for new_proj in new_proj_candidates:
                    try:
                        new_proj.first.wait_for(state="visible", timeout=8000)
                    except Exception:
                        pass
                    if new_proj.count() > 0 and new_proj.first.is_visible():
                        new_proj.first.click()
                        clicked = True
                        break
                if clicked:
                    log("Clicked 'New project' — waiting for editor to load...")
                    try:
                        page.wait_for_load_state("networkidle", timeout=10_000)
                    except Exception:
                        pass
                    page.wait_for_timeout(3000)
                    log(f"Editor URL: {page.url}")
                else:
                    log("No 'New project' button found — may already be in editor")
            except Exception as e:
                log(f"WARNING: Could not click 'New project' ({e}) — proceeding")
        else:
            log(f"Already on project page: {page.url}")

        page.wait_for_timeout(2000)  # Allow SPA to finish rendering

        # Dismiss any changelog / "What's New" popup that blocks the editor
        try:
            if page.locator("iframe[src*='changelog']").count() > 0:
                log("Changelog popup detected — dismissing with Escape...")
                page.keyboard.press("Escape")
                page.wait_for_timeout(800)
        except Exception:
            pass

        # ── Find prompt input (once — reused across retries) ───────────────────
        log("Looking for prompt input...")
        prompt_input = find_prompt_input(page)
        if prompt_input is None:
            log("ERROR: Could not find the prompt input field.")
            log("The UI may have changed. Run scripts/discover_selectors.py to debug.")
            context.close()
            browser.close()
            sys.exit(1)

        # ── Video-specific settings (mode switch + model + variation count) ──────
        if args.type == "video":
            select_video_model(page, "Veo 3.1 - Lite")
            set_variation_count(page, 1)
            # Dismiss any dropdown/panel left open by the settings interactions
            page.keyboard.press("Escape")
            page.wait_for_timeout(500)

        # ── Install network interceptor before generating ──────────────────────
        intercepted: list[str] = []
        video_urls:  list[str] = []
        install_response_interceptor(page, intercepted, args.type, video_urls)


        # ── Generate with retry on quota errors ────────────────────────────────
        result_url = None
        for attempt in range(1, args.retries + 2):  # +2: retries=3 → attempts 1..4
            if attempt > 1:
                log(f"Retry {attempt - 1}/{args.retries} after {args.retry_delay}s...")
                time.sleep(args.retry_delay)
                intercepted.clear()
                video_urls.clear()

            prompt_input.click()
            # Clear any existing content then type the prompt.
            # contenteditable divs need select-all + type rather than fill().
            page.keyboard.press("Meta+a")  # macOS select-all
            page.keyboard.press("Control+a")  # Linux/Windows fallback
            page.keyboard.type(args.prompt)
            page.wait_for_timeout(500)
            entered = (prompt_input.inner_text() or "").strip()
            log(f"Prompt verified: '{entered[:60]}{'...' if len(entered) > 60 else ''}'")
            if not entered:
                log("WARNING: Prompt field appears empty — text input may not have worked")

            log("Looking for generate button...")
            gen_button = find_generate_button(page)
            if gen_button is None:
                log("ERROR: Could not find the generate/create button.")
                log("The UI may have changed. Run scripts/discover_selectors.py to debug.")
                context.close()
                browser.close()
                sys.exit(1)

            pre_urls = snapshot_image_urls(page) if args.type == "image" else None
            log(f"Generate button text: '{(gen_button.text_content() or '').strip()[:60]}'")
            gen_button.click()
            log(f"Generate clicked (attempt {attempt}). Waiting for result...")

            result_url = wait_for_result(page, args.type, args.timeout,
                                         pre_urls=pre_urls, intercepted=intercepted,
                                         video_urls=video_urls)

            if result_url == "QUOTA_ERROR":
                if attempt <= args.retries:
                    log(f"Quota limit hit (error 253). Waiting {args.retry_delay}s before retry...")
                    result_url = None
                    continue
                else:
                    log("Quota limit hit and retries exhausted.")
                    result_url = None
                    break
            break  # success or non-quota failure

        if result_url is None:
            log("Generation failed or timed out.")
            context.close()
            browser.close()
            sys.exit(1)

        log(f"Result found: {result_url[:80]}...")

        # ── Download ───────────────────────────────────────────────────────────
        out_path = download_result(result_url, args.type, args.out, page, args.name)

        context.close()
        browser.close()

    print(f"\n✅ Saved to: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    main()
